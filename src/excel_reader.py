from io import BytesIO
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from .column_mapping import key


def detect_header_row(ws, expected_terms, max_rows=40):
    wanted = {key(x) for x in expected_terms}
    best = (0, 1)
    for row_no in range(1, min(ws.max_row, max_rows) + 1):
        values = {key(c.value) for c in ws[row_no] if c.value not in (None, "")}
        score = len(values & wanted)
        if score > best[0]:
            best = (score, row_no)
    if best[0] == 0:
        raise ValueError("Could not detect a table header row.")
    return best[1]


def inspect_workbook(source, expected_terms):
    wb = load_workbook(source, read_only=False, data_only=True)
    candidates = []
    for ws in wb.worksheets:
        try:
            header = detect_header_row(ws, expected_terms)
            values = [c.value for c in ws[header]]
            score = len({key(v) for v in values if v} & {key(x) for x in expected_terms})
            candidates.append((score, ws.title, header, values))
        except ValueError:
            continue
    if not candidates:
        raise ValueError("No worksheet with recognizable headers was found.")
    return max(candidates, key=lambda x: x[0])


def read_table(source, expected_terms):
    score, sheet, header, _ = inspect_workbook(source, expected_terms)
    df = pd.read_excel(source, sheet_name=sheet, header=header - 1, dtype=object)
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    df.columns = [str(c).strip() for c in df.columns]
    return df.reset_index(drop=True), {"sheet": sheet, "header_row": header, "score": score}

