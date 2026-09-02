from io import BytesIO
import json
from collections import Counter, defaultdict

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DIRECT_EXPENSE_COLUMNS = [
    "Route expense", "Bill discounting", "Salary", "Driver's salary", "Rent",
    "Office & General expenses", "Conveyance", "EMI", "Insurance", "Vehicle Tax",
    "Repair and maintenance", "Interest", "Extra Expense",
]
REPORT_EXPENSE_COLUMNS = [*DIRECT_EXPENSE_COLUMNS, "Passing expense"]
BRANCH_PNL_COLUMNS = [
    "Branch", "Revenue-Own", "Revenue OS", "Total Revenue", "Transporter Freight",
    "Extra Exp", "Passing Exp", "Bill Discounting", "UPI", "Salary", "Rent",
    "Office Expense", "Conveyance", "EMI", "Ins/Tax", "R & M", "Toll",
    "Driver's Salary", "Diesel", "Interest", "Expense", "Profit",
]


def pnl_summary(trip_rows, expense_rows):
    revenue = sum(float(row.get("revenue") or 0) for row in trip_rows)
    transporter = sum(float(row.get("transporter_freight") or 0) for row in trip_rows)
    branches = sorted({str(row.get("branch") or "").strip() for row in trip_rows} - {""}, key=str.casefold)
    toll = sum(_amount(row.get("dtr_data", {}), "Toll Expense") for row in trip_rows)
    diesel = sum(_amount(row, "diesel_advance") for row in trip_rows)
    categories = {name: 0.0 for name in REPORT_EXPENSE_COLUMNS}
    for row in expense_rows:
        for name, value in row.get("categories", {}).items():
            if name in categories:
                categories[name] += float(value or 0)
    direct_total = sum(categories.values())
    gross = revenue - transporter
    return [
        {"Particular": "Branch", "Amount": ", ".join(branches) or "—"},
        {"Particular": "Revenue", "Amount": revenue},
        {"Particular": "Transporter Freight", "Amount": -transporter},
        {"Particular": "Gross Contribution", "Amount": gross},
        {"Particular": "Toll charges", "Amount": -toll},
        {"Particular": "Diesel", "Amount": -diesel},
        *({"Particular": name, "Amount": -amount} for name, amount in categories.items()),
        {"Particular": "Additional expenses", "Amount": -direct_total},
        {"Particular": "Total Direct Expenses", "Amount": -direct_total},
        {"Particular": "Net Profit / (Loss)", "Amount": gross - toll - diesel - direct_total},
    ]


def _amount(row, field):
    if isinstance(row, str):
        try:
            row = json.loads(row)
        except (TypeError, ValueError):
            row = {}
    if not isinstance(row, dict):
        row = {}
    try:
        return float(row.get(field) or 0)
    except (TypeError, ValueError):
        return 0.0


def _category_total(expense_rows, *names):
    return sum(
        _amount(row.get("categories", {}), name)
        for row in expense_rows
        for name in names
    )


def _trip_amount(row, field, dtr_field):
    stored = _amount(row, field)
    return stored if stored else _amount(row.get("dtr_data", {}), dtr_field)


def branch_pnl_summary(trip_rows, expense_rows):
    """Build the horizontal, branch-wise P&L used by the Both report."""
    vehicle_branches = defaultdict(Counter)
    for row in trip_rows:
        vehicle = str(row.get("vehicle_number") or "").strip().casefold()
        branch = str(row.get("branch") or "").strip()
        if vehicle and branch:
            vehicle_branches[vehicle][branch] += 1

    expenses_by_branch = defaultdict(list)
    for row in expense_rows:
        branch = str(row.get("branch") or "").strip()
        if not branch:
            vehicle = str(row.get("vehicle_number") or "").strip().casefold()
            if vehicle_branches.get(vehicle):
                branch = vehicle_branches[vehicle].most_common(1)[0][0]
        expenses_by_branch[branch or "Not specified"].append(row)

    trips_by_branch = defaultdict(list)
    for row in trip_rows:
        branch = str(row.get("branch") or "").strip() or "Not specified"
        trips_by_branch[branch].append(row)
    branches = sorted(set(trips_by_branch) | set(expenses_by_branch), key=str.casefold)

    rows = []
    for branch in branches:
        trips = trips_by_branch[branch]
        direct = expenses_by_branch[branch]
        own = [row for row in trips if str(row.get("ownership_type") or "").strip().lower().startswith("own")]
        outside = [row for row in trips if str(row.get("ownership_type") or "").strip().lower().startswith("outside")]
        categories = {name: _category_total(direct, name) for name in REPORT_EXPENSE_COLUMNS}
        revenue_own = sum(_amount(row, "revenue") for row in own)
        revenue_outside = sum(_amount(row, "revenue") for row in outside)
        transporter = sum(_amount(row, "transporter_freight") for row in outside)
        upi = sum(_trip_amount(row, "upi", "UPI") for row in own)
        diesel = sum(_trip_amount(row, "diesel_advance", "Diesel Adv.") for row in own)
        toll = sum(_amount(row.get("dtr_data", {}), "Toll Expense") for row in own)
        repairs = categories["Repair and maintenance"] + sum(
            _amount(row.get("dtr_data", {}), "Repairs & Maintenance") for row in own
        )
        row = {
            "Branch": branch,
            "Revenue-Own": revenue_own,
            "Revenue OS": revenue_outside,
            "Total Revenue": revenue_own + revenue_outside,
            "Transporter Freight": transporter,
            "Extra Exp": categories["Route expense"] + categories["Extra Expense"],
            "Passing Exp": categories["Passing expense"],
            "Bill Discounting": categories["Bill discounting"],
            "UPI": upi,
            "Salary": categories["Salary"],
            "Rent": categories["Rent"],
            "Office Expense": categories["Office & General expenses"],
            "Conveyance": categories["Conveyance"],
            "EMI": categories["EMI"],
            "Ins/Tax": categories["Insurance"] + categories["Vehicle Tax"],
            "R & M": repairs,
            "Toll": toll,
            "Driver's Salary": categories["Driver's salary"],
            "Diesel": diesel,
            "Interest": categories["Interest"],
        }
        row["Expense"] = sum(row[column] for column in BRANCH_PNL_COLUMNS[4:-2])
        row["Profit"] = row["Total Revenue"] - row["Expense"]
        rows.append(row)

    total = {"Branch": "Total"}
    for column in BRANCH_PNL_COLUMNS[1:]:
        total[column] = sum(row[column] for row in rows)
    return [*rows, total] if rows else []


def vehicle_pnl_summary(trip_rows, expense_rows, ownership):
    """Return the compact operations P&L requested for own/outside vehicles."""
    ownership = str(ownership or "Both")
    own_trips = [row for row in trip_rows if str(row.get("ownership_type") or "").strip().lower().startswith("own")]
    outside_trips = [row for row in trip_rows if str(row.get("ownership_type") or "").strip().lower().startswith("outside")]
    own_vehicles = {str(row.get("vehicle_number") or "").strip().casefold() for row in own_trips} - {""}
    outside_vehicles = {str(row.get("vehicle_number") or "").strip().casefold() for row in outside_trips} - {""}
    if own_trips and outside_trips:
        own_expenses = [row for row in expense_rows if str(row.get("vehicle_number") or "").strip().casefold() in own_vehicles]
        outside_expenses = [row for row in expense_rows if str(row.get("vehicle_number") or "").strip().casefold() in outside_vehicles]
    else:
        own_expenses = outside_expenses = expense_rows

    def own_rows():
        revenue = sum(_amount(row, "revenue") for row in own_trips)
        route = sum(_trip_amount(row, "upi", "UPI") for row in own_trips)
        toll = sum(_amount(row.get("dtr_data", {}), "Toll Expense") for row in own_trips)
        diesel = sum(_trip_amount(row, "diesel_advance", "Diesel Adv.") for row in own_trips)
        driver_salary = _category_total(own_expenses, "Driver's salary")
        emi = _category_total(own_expenses, "EMI")
        insurance = _category_total(own_expenses, "Insurance")
        vehicle_tax = _category_total(own_expenses, "Vehicle Tax")
        repairs = sum(_amount(row.get("dtr_data", {}), "Repairs & Maintenance") for row in own_trips)
        repairs += _category_total(own_expenses, "Repair and maintenance")
        expenses = route + toll + diesel + driver_salary + emi + insurance + vehicle_tax + repairs
        return [
            {"Particular": "Revenue freight", "Amount": revenue},
            {"Particular": "Route expenses (UPI)", "Amount": -route},
            {"Particular": "Toll charges", "Amount": -toll},
            {"Particular": "Diesel amount", "Amount": -diesel},
            {"Particular": "Driver's salary", "Amount": -driver_salary},
            {"Particular": "EMI", "Amount": -emi},
            {"Particular": "Insurance", "Amount": -insurance},
            {"Particular": "Vehicle Tax", "Amount": -vehicle_tax},
            {"Particular": "Repair and maintenance", "Amount": -repairs},
            {"Particular": "Net Profit / (Loss)", "Amount": revenue - expenses},
        ]

    def outside_rows():
        revenue = sum(_amount(row, "revenue") for row in outside_trips)
        transporter = sum(_amount(row, "transporter_freight") for row in outside_trips)
        additional = sum(_amount(row, "amount") for row in outside_expenses)
        return [
            {"Particular": "Revenue", "Amount": revenue},
            {"Particular": "Transporter Freight", "Amount": -transporter},
            {"Particular": "Additional expenses", "Amount": -additional},
            {"Particular": "Net Profit / (Loss)", "Amount": revenue - transporter - additional},
        ]

    own = own_rows()
    outside = outside_rows()
    if ownership == "Own":
        return own
    if ownership == "Outside":
        return outside
    return pnl_summary(trip_rows, expense_rows)


def branch_vehicle_pnl_summary(trip_rows, expense_rows, ownership):
    """Transpose the existing Own/Outside P&L fields into branch-wise rows."""
    selected = [
        row for row in trip_rows
        if str(row.get("ownership_type") or "").strip().lower().startswith(ownership.lower())
    ]
    trips_by_branch = defaultdict(list)
    vehicle_branch = {}
    for row in selected:
        branch = str(row.get("branch") or "").strip() or "Not specified"
        trips_by_branch[branch].append(row)
        vehicle = str(row.get("vehicle_number") or "").strip().casefold()
        if vehicle:
            vehicle_branch[vehicle] = branch
    expenses_by_branch = defaultdict(list)
    for row in expense_rows:
        vehicle = str(row.get("vehicle_number") or "").strip().casefold()
        branch = str(row.get("branch") or "").strip() or vehicle_branch.get(vehicle)
        if branch in trips_by_branch:
            expenses_by_branch[branch].append(row)
    rows = []
    for branch in sorted(trips_by_branch, key=str.casefold):
        vertical = vehicle_pnl_summary(trips_by_branch[branch], expenses_by_branch[branch], ownership)
        rows.append({"Branch": branch, **{item["Particular"]: item["Amount"] for item in vertical}})
    if rows:
        total = {"Branch": "Total"}
        for column in rows[0]:
            if column != "Branch":
                total[column] = sum(float(row.get(column) or 0) for row in rows)
        rows.append(total)
    return rows


def export_pnl(trip_rows, expense_rows, start_date, end_date, ownership=None):
    rows = branch_vehicle_pnl_summary(trip_rows, expense_rows, ownership) if ownership in {"Own", "Outside"} else branch_pnl_summary(trip_rows, expense_rows)
    frame = pd.DataFrame(rows)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name="P&L", startrow=2)
        ws = writer.book["P&L"]
        ws["A1"] = f"Profit & Loss | {start_date:%d-%m-%Y} to {end_date:%d-%m-%Y}"
        ws["A1"].font = Font(size=14, bold=True, color="17324D")
        final_column = max(2, len(frame.columns))
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=final_column)
        fill = PatternFill("solid", fgColor="0F766E")
        for cell in ws[3]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(1)].width = 20
        for column_index in range(2, final_column + 1):
            ws.column_dimensions[get_column_letter(column_index)].width = 18
            for row_index in range(4, ws.max_row + 1):
                ws.cell(row=row_index, column=column_index).number_format = '₹#,##0.00;[Red]-₹#,##0.00'
    return output.getvalue()
