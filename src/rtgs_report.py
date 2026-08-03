import datetime as dt
import json
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RTGS_COLUMNS = [
    "File_Sequence_Num", "Pymt_Prod_Type_Code", "Pymt_Mode", "Debit_Acct_no",
    "Beneficiary Name", "Beneficiary Account No", "Bene_IFSC_Code", "Amount",
    "Debit narration", "Credit narration", "Mobile Numder", "Email id", "Remark",
    "Pymt_Date", "Reference_no", "Addl_Info1", "Addl_Info2", "Addl_Info3",
    "Addl_Info4", "Addl_Info5", "STATUS", "Current Step", "File name",
    "Rejected by", "Rejection Reason", "Acct_Debit_date", "Customer Ref No", "UTR NO",
]


def rows_to_rtgs(rows):
    records = []
    for row in rows:
        raw = row.get("rtgs_data") or "{}"
        data = json.loads(raw) if isinstance(raw, str) else dict(raw)
        record = {column: data.get(column, "") for column in RTGS_COLUMNS}
        record["Beneficiary Name"] = record["Beneficiary Name"] or row.get("beneficiary_name", "")
        record["Amount"] = record["Amount"] if record["Amount"] not in (None, "") else row.get("amount", "")
        record["Pymt_Date"] = record["Pymt_Date"] or row.get("trip_date", "")
        records.append(record)
    return pd.DataFrame(records, columns=RTGS_COLUMNS)


def export_rtgs(df):
    output = BytesIO()
    safe = df.reindex(columns=RTGS_COLUMNS).copy()
    for column in ("File_Sequence_Num", "Debit_Acct_no", "Beneficiary Account No", "Mobile Numder", "Customer Ref No", "UTR NO"):
        safe[column] = safe[column].fillna("").astype(str)
    for column in ("Pymt_Date", "Acct_Debit_date"):
        safe[column] = pd.to_datetime(safe[column], errors="coerce").dt.date
    with pd.ExcelWriter(output, engine="openpyxl", date_format="DD-MM-YYYY") as writer:
        safe.to_excel(writer, index=False, sheet_name="Sheet0")
        ws = writer.book["Sheet0"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for index, name in enumerate(RTGS_COLUMNS, 1):
            letter = get_column_letter(index)
            lengths = [len(name)] + [len(str(value)) if value is not None else 0 for value in safe[name]]
            ws.column_dimensions[letter].width = min(max(max(lengths, default=0) + 2, 12), 35)
            for cell in ws[letter][1:]:
                if name in ("Pymt_Date", "Acct_Debit_date") and cell.value is not None:
                    cell.number_format = "dd-mm-yyyy"
                elif name in ("File_Sequence_Num", "Debit_Acct_no", "Beneficiary Account No", "Mobile Numder", "Customer Ref No", "UTR NO"):
                    cell.number_format = "@"
    return output.getvalue()
