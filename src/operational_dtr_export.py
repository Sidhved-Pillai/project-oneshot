from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .ai_intake import DTR_REVIEW_COLUMNS


OPERATIONAL_DTR_COLUMNS = [column for column in DTR_REVIEW_COLUMNS if column != "Review Notes"]
EXPORT_NAMES = {"Compnay Name": "Company Name", "UPI": "UPI "}


def export_operational_dtr(df):
    output = BytesIO()
    safe = df.reindex(columns=OPERATIONAL_DTR_COLUMNS).copy()
    safe = safe.rename(columns=EXPORT_NAMES)
    date_columns = ["Date", "Received Date"]
    for column in date_columns:
        safe[column] = pd.to_datetime(safe[column], errors="coerce").dt.date
    with pd.ExcelWriter(output, engine="openpyxl", date_format="DD-MM-YYYY") as writer:
        safe.to_excel(writer, index=False, sheet_name="DTR")
        ws = writer.book["DTR"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        text_columns = {"Vehicle No.", "LR No.", "Invoice No.", "Bill No."}
        for index, name in enumerate(safe.columns, 1):
            letter = get_column_letter(index)
            lengths = [len(name)] + [len(str(value)) if value is not None else 0 for value in safe[name]]
            ws.column_dimensions[letter].width = min(max(max(lengths, default=0) + 2, 11), 35)
            for cell in ws[letter][1:]:
                if name in date_columns and cell.value is not None:
                    cell.number_format = "dd-mm-yyyy"
                elif name in text_columns:
                    cell.number_format = "@"
    return output.getvalue()
