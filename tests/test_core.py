from io import BytesIO
import datetime as dt
import pandas as pd
from openpyxl import Workbook, load_workbook
import pytest
from src.excel_reader import detect_header_row
from src.column_mapping import resolve_columns, DTR_ALIASES, CONSOLIDATED_ALIASES
from src.remark_classifier import classify_remark
from src.remark_parser import normalize_vehicle, last_four, normalize_vehicle_type, parse_remark
from src.vehicle_matcher import find_vehicle, resolve_vehicle, choose_vehicle_type
from src.master_builder import build_vehicle_master, build_company_master, build_beneficiary_master
from src.dtr_generator import generate_dtr, final_review_rows, DTR_COLUMNS, FINANCIAL_COLUMNS
from src.excel_exporter import export_dtr
from src.gemini_parser import parse_with_gemini
from src.historical_suggester import HistoricalSuggester
from src.entry_finance import advance_summary, financial_values
from src.request_store import RequestStore, rows_to_dtr
from src.rtgs_report import RTGS_COLUMNS, export_rtgs, normalize_rtgs_records, rows_to_rtgs
from src.operational_dtr_export import OPERATIONAL_DTR_COLUMNS, export_operational_dtr
from src.pnl_report import BRANCH_PNL_COLUMNS, DIRECT_EXPENSE_COLUMNS, branch_pnl_summary, branch_vehicle_pnl_summary, export_pnl, pnl_summary, vehicle_pnl_summary
from src.text_normalization import canonical_company, canonical_location, canonical_vehicle_capacity, plain_remark
from src.business_memory import build_business_memory, recall
from src.workflow_ai import convert_rtgs_to_dtr as workflow_convert_rtgs_to_dtr
from src.workflow_pnl import branch_vehicle_pnl_summary as workflow_branch_vehicle_pnl_summary
from src.workflow_store import RequestStore as WorkflowRequestStore
from src.ai_intake import DTRIntakeResult, DTRIntakeRow, _model_unavailable, _prompt, extract_intake, result_to_records


def vehicle_master():
    return pd.DataFrame([
        {"Vehicle No.": "MH14JL9818", "Last 4 Digits": "9818", "Vehicle Type": "10MT", "Ownership Type": "Outside Vehicle", "Transporter Name": "Vehicle Carrier", "Vehicle Status": "Active"},
        {"Vehicle No.": "MH12AB9818", "Last 4 Digits": "9818", "Vehicle Type": "12MT", "Ownership Type": "Own Vehicle", "Transporter Name": "", "Vehicle Status": "Active"},
        {"Vehicle No.": "MH04AA8172", "Last 4 Digits": "8172", "Vehicle Type": "21MT", "Ownership Type": "Outside Vehicle", "Transporter Name": "Master Carrier", "Vehicle Status": "Active"},
        {"Vehicle No.": "MH01AB0272", "Last 4 Digits": "0272", "Vehicle Type": "20MT", "Ownership Type": "Outside Vehicle", "Transporter Name": "Zero Carrier", "Vehicle Status": "Active"},
        {"Vehicle No.": "MH99XX0001", "Last 4 Digits": "0001", "Vehicle Type": "10MT", "Ownership Type": "Outside Vehicle", "Transporter Name": "", "Vehicle Status": "Inactive"},
    ])


def beneficiary_master():
    return pd.DataFrame([{"Beneficiary Account No.": "001234", "Beneficiary Name": "Demo", "Transporter Name": "Account Carrier"}])


def source(remarks):
    return pd.DataFrame({"Remark": remarks, "Beneficiary Name": ["Demo"] * len(remarks),
                         "Beneficiary Account No": ["001234"] * len(remarks)})


def test_header_detection_with_blank_rows():
    ws = Workbook().active; ws.append([]); ws.append(["report title"]); ws.append(["Remark", "Beneficiary Name"])
    assert detect_header_row(ws, ["Remark", "Beneficiary Name"]) == 3


def test_column_mappings():
    assert resolve_columns(["Compnay Name", "Own/Outside Veh."], DTR_ALIASES)["Own/Outside Vehicle"] == "Own/Outside Veh."
    assert resolve_columns(["Remarks", "Beneficiary Name"], CONSOLIDATED_ALIASES)["Remark"] == "Remarks"


def test_vehicle_normalization_and_last_four():
    assert normalize_vehicle("mh 14-jl 9818") == "MH14JL9818"
    assert last_four("MH14JL9818") == "9818"


def test_operational_text_normalization_is_conservative():
    assert canonical_company("saint gobin") == "Saint-Gobain India Private Limited"
    assert canonical_company("saint gobain gyproc") == "Saint-Gobain India Private Limited - Gyproc"
    assert canonical_location("talegoan") == "Talegaon"
    assert canonical_location("PUNE") == "Pune"
    assert canonical_vehicle_capacity("03 tons") == "3 MT"
    assert canonical_vehicle_capacity("10 ton") == "10 MT"
    assert canonical_vehicle_capacity("12mt") == "12 MT"
    assert plain_remark("1234", "Pune-to-Wada", "10 MT", "TA") == "1234 Pune to Wada 10 MT TA"


def test_duplicate_suffix_and_inactive_filtering():
    assert len(find_vehicle("9818", vehicle_master())) == 2
    assert find_vehicle("0001", vehicle_master()) == []


def test_unique_suffix_resolves_to_full_vehicle_and_master_fields():
    number, match, choices = resolve_vehicle(["0272"], vehicle_master())
    assert number == "MH01AB0272" and match["Vehicle Type"] == "20MT" and choices == ["MH01AB0272"]


def test_full_registration_normalizes_and_matches_directly():
    number, match, choices = resolve_vehicle(["mh 04-aa 8172"], vehicle_master())
    assert number == "MH04AA8172" and match["Ownership Type"] == "Outside Vehicle"


def test_unmatched_suffix_remains_unresolved_and_leading_zero_survives():
    number, match, choices = resolve_vehicle(["0099"], vehicle_master())
    assert number == "0099" and not match and choices == []


def test_multiple_identifiers_preserved_without_automatic_choice():
    parsed = parse_remark("6789 7890 Shriwal to Bhiwandi 12MT 29 30 07 2026 TA")
    assert parsed["vehicle_identifiers"] == ["6789", "7890"] and parsed["vehicle_identifier"] == ""


def test_vehicle_conflicts_are_not_silently_resolved():
    a = pd.DataFrame({"Vehicle No.":["MH14JL9818"],"Vehicle Type":["10MT"],"Own/Outside Veh.":["Outside Vehicle"],"Transporter Name":["A"]})
    b = a.copy(); b.loc[0,"Vehicle Type"] = "12MT"
    master, conflicts = build_vehicle_master([a,b])
    assert master.iloc[0]["Vehicle Type"] == "" and len(conflicts) == 1


@pytest.mark.parametrize("remark,master,expected,conflict", [
    ("25MT", "", "25 MT", False), ("", "21MT", "21 MT", False),
    ("25MT", "25MT", "25 MT", False), ("25MT", "21MT", "25 MT", True),
])
def test_vehicle_type_precedence(remark, master, expected, conflict):
    assert choose_vehicle_type(remark, master) == (expected, conflict)


def test_company_branch_deduplication():
    df = pd.DataFrame({"Compnay Name":["SG","SG"],"Branch":["Pune","Pune"]})
    assert len(build_company_master([df])) == 1


def test_beneficiary_accounts_preserved_as_strings():
    df = pd.DataFrame({"Beneficiary Account No":["001234","001234"],"Beneficiary Name":["Demo","Demo"]})
    master, _ = build_beneficiary_master(df)
    assert master.iloc[0]["Beneficiary Account No."] == "001234" and len(master) == 1


def test_salary_is_confirmed_non_trip_and_number_not_vehicle():
    result = classify_remark("Driver Salary for June 2026 8595 Laxman", ["8595"])
    assert (result.classification, result.reason) == ("Confirmed Non-trip", "Driver Salary")
    assert parse_remark("Driver Salary for June 2026 8595 Laxman")["vehicle_identifiers"] == []


@pytest.mark.parametrize("remark,expected", [
    ("7348 Wagholi to Kamshet 07 08 2026 TA", "Confirmed Trip"),
    ("9416 Jhagadia to Indore 09MT 31 07 2026 TA", "Confirmed Trip"),
    ("6765 Talegaon to Pimpri 30 07 2026 TA", "Confirmed Trip"),
    ("6305 Bhiwandi to Pune Balance Payment 2LR TP", "Potential Trip"),
    ("9818 29 07 2026 to 31 07 2026 3Trp TA", "Potential Trip"),
    ("4094 25 07 2026 to 28 07 2026 2Trp TA", "Potential Trip"),
    ("Balance Payment as per ledger TP", "Potential Trip"),
    ("Amount paid as Transporter Payment Thru UPI", "Potential Trip"),
    ("Vasai Local Transporter Payment", "Potential Trip"),
])
def test_classification_regressions(remark, expected):
    assert classify_remark(remark).classification == expected


@pytest.mark.parametrize("remark,vehicle,origin,destination,kind,date", [
    ("7348 Wagholi to Kamshet 07 08 2026 TA", "7348", "Wagholi", "Kamshet", "", "07-08-2026"),
    ("9416 Jhagadia to Indore 09MT 31 07 2026 TA", "9416", "Jhagadia", "Indore", "9 MT", "31-07-2026"),
    ("6765 Talegaon to Pimpri 30 07 2026 TA", "6765", "Talegaon", "Pimpri", "", "30-07-2026"),
    ("6305 Bhiwandi to Pune Balance Payment 2LR TP", "6305", "Bhiwandi", "Pune", "", None),
    ("3946 Talegaon to Hinjewadi 5k card diesal 25MT 26 07 2026 TA", "3946", "Talegaon", "Hinjewadi", "25 MT", "26-07-2026"),
    ("0272 Wagholi to Kamshet 20MT 27 07 2026 TA", "0272", "Wagholi", "Kamshet", "20 MT", "27-07-2026"),
])
def test_actual_style_route_parsing(remark, vehicle, origin, destination, kind, date):
    parsed = parse_remark(remark)
    assert parsed["vehicle_identifiers"] == [vehicle]
    assert parsed["from_location"] == origin and parsed["to_location"] == destination
    assert parsed["vehicle_type"] == kind
    assert (parsed["date"].strftime("%d-%m-%Y") if parsed["date"] is not None else None) == date


def test_unusual_adjacent_days_best_effort():
    parsed = parse_remark("6789 7890 Shriwal to Bhiwandi 12MT 29 30 07 2026 TA")
    assert parsed["from_location"] == "Shriwal" and parsed["to_location"] == "Bhiwandi"
    assert parsed["vehicle_type"] == "12 MT" and parsed["date"].strftime("%d-%m-%Y") == "30-07-2026"


@pytest.mark.parametrize("value", ["30 07 2026", "30/07/2026", "30.07.26"])
def test_all_supported_remark_date_formats(value):
    assert parse_remark(f"9818 Pune to Mumbai 10MT {value} TA")["date"].strftime("%d-%m-%Y") == "30-07-2026"


def test_pipe_format_extracts_company_and_fields():
    parsed = parse_remark("9818 | SG | Talegaon to Bhiwandi | 10MT | 30 07 2026 | INV 0012345 | TA")
    assert parsed["vehicle_identifiers"] == ["9818"]
    assert parsed["company_name"] == "SG" and parsed["from_location"] == "Talegaon" and parsed["to_location"] == "Bhiwandi"
    assert parsed["vehicle_type"] == "10 MT" and parsed["date"].strftime("%d-%m-%Y") == "30-07-2026"
    assert parsed["invoice_number"] == "0012345"


@pytest.mark.parametrize("remark,date", [
    ("9818 29 07 2026 to 31 07 2026 3Trp TA", "31-07-2026"),
    ("4094 25 07 2026 to 28 07 2026 2Trp TA", "28-07-2026"),
])
def test_date_range_uses_end_date_one_row(remark, date):
    confirmed, potential, _ = generate_dtr(source([remark]), vehicle_master(), beneficiary_master())
    assert confirmed.empty and len(potential) == 1
    assert potential.iloc[0]["Date"].strftime("%d-%m-%Y") == date
    assert potential.iloc[0]["From"] == "" and potential.iloc[0]["To"] == ""


@pytest.mark.parametrize("marker", ["INV", "Invoice", "Invoice No", "Bill No"])
def test_invoice_markers_and_leading_zero(marker):
    assert parse_remark(f"1234 Pune to Mumbai 10MT {marker} 00127")["invoice_number"] == "00127"


def test_generator_account_priority_and_remark_type_conflict():
    confirmed, potential, non_trip = generate_dtr(source(["8172 Jhagadia to Indore 25MT 31 07 2026 TA"]), vehicle_master(), beneficiary_master())
    row = confirmed.iloc[0]
    assert row["Vehicle No."] == "MH04AA8172" and row["Vehicle Type"] == "25 MT" and bool(row["_needs_review"]) and bool(row["_type_conflict"])
    assert row["Transporter Name"] == "Account Carrier"


def test_beneficiary_always_comes_from_uploaded_report_without_master():
    empty_beneficiaries = pd.DataFrame(columns=["Beneficiary Account No.", "Beneficiary Name", "Transporter Name"])
    confirmed, _, _ = generate_dtr(source(["7348 Wagholi to Kamshet 10MT 07 08 2026 TA"]), vehicle_master(), empty_beneficiaries)
    assert confirmed.iloc[0]["Benificiary Name"] == "Demo" and confirmed.iloc[0]["Transporter Name"] == ""


def test_payment_date_never_fills_missing_remark_date():
    frame = source(["7348 Wagholi to Kamshet 10MT TA"])
    frame["Pymt_Date"] = ["30-07-2026"]
    confirmed, _, _ = generate_dtr(frame, vehicle_master(), beneficiary_master())
    assert pd.isna(confirmed.iloc[0]["Date"])


def test_historical_suggestions_are_conservative_and_explicit_company_wins():
    suggester = HistoricalSuggester({
        "routes": {"jhagadia|indore": {"company": "SG", "branch": "Vadodara"}},
        "route_companies": {"talegaon|bhiwandi|sg": {"branch": "Pune"}},
    })
    assert suggester.suggest("Jhagadia", "Indore") == ("SG", "Vadodara")
    assert suggester.suggest("Talegaon", "Bhiwandi", "SG") == ("SG", "Pune")
    assert suggester.suggest("New", "Route") == ("", "")


def test_every_input_is_accounted_for_and_potential_not_auto_included():
    remarks = ["7348 Wagholi to Kamshet 07 08 2026 TA", "Balance Payment as per ledger TP", "Driver Salary for June 2026 8595 Laxman"]
    confirmed, potential, non_trip = generate_dtr(source(remarks), vehicle_master(), beneficiary_master())
    assert len(confirmed) + len(potential) + len(non_trip) == len(remarks)
    assert not potential.iloc[0]["Include in DTR"]
    assert len(final_review_rows(confirmed, potential)) == 1
    potential.loc[:, "Include in DTR"] = True
    assert len(final_review_rows(confirmed, potential)) == 2


def test_exact_columns_blank_financials_and_export_date_format_width():
    frame = pd.DataFrame([{**{c:"" for c in DTR_COLUMNS}, "Sr No.":1, "Date":pd.Timestamp("2026-07-30"),
                           "Vehicle No.":"MH01AB0272", "Invoice No.":"00127", "Revenue":"999"}])
    data = export_dtr(frame); ws = load_workbook(BytesIO(data))["DTR"]
    assert [c.value for c in ws[1]] == DTR_COLUMNS and ws.freeze_panes == "A2" and ws.auto_filter.ref
    assert isinstance(ws["D2"].value, (dt.datetime, dt.date)) and ws["D2"].number_format == "dd-mm-yyyy"
    assert 13 <= ws.column_dimensions["D"].width <= 15
    assert ws["E2"].number_format == "@" and ws["I2"].number_format == "@" and ws["I2"].value == "00127"
    assert all(ws.cell(2, DTR_COLUMNS.index(c)+1).value is None for c in FINANCIAL_COLUMNS)
    assert ws.max_column == 22


def test_missing_required_column():
    with pytest.raises(ValueError, match="Missing required"):
        generate_dtr(pd.DataFrame({"Remark":["trip"]}), vehicle_master(), beneficiary_master())


def test_gemini_unavailable(monkeypatch):
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)
    assert parse_with_gemini("ambiguous") is None


def test_financial_mapping_and_persistent_request_roundtrip(tmp_path):
    finance = financial_values("Trip Advance", "UPI", 1250, 12.5)
    assert finance["upi"] == 1250 and finance["total_advance"] == 1250
    store = RequestStore(f"sqlite:///{tmp_path / 'requests.db'}")
    number = store.create({
        "trip_date": dt.date(2026, 8, 3), "vehicle_number": "MH14JL9818",
        "invoice_number": "INV-7", "company_name": "Demo Co", "expense_type": "Trip Advance",
        "amount": 1250, "payment_mode": "UPI", **finance, "status": "Submitted",
        "source_filename": "proof.jpg", "source_mime_type": "image/jpeg", "source_image": b"proof",
    })
    assert number.startswith("REQ-202608-")
    reopened = RequestStore(f"sqlite:///{tmp_path / 'requests.db'}")
    rows = reopened.list(dt.date(2026, 8, 1), dt.date(2026, 8, 31))
    assert len(rows) == 1 and rows[0]["source_image"] == b"proof"
    dtr = rows_to_dtr(rows)
    assert dtr.iloc[0]["UPI"] == 1250 and dtr.iloc[0]["Invoice No."] == "INV-7"


def test_advance_summary_uses_all_modes_and_preserves_overpayment():
    normal = advance_summary(38000, 18000, 2000, 3000, 1000)
    assert normal["total_advance"] == 24000
    assert normal["balance_payable"] == 14000
    with_billtee = advance_summary(38000, 18000, 2000, 3000, 1000, 1000)
    assert with_billtee["total_advance"] == 25000
    assert with_billtee["balance_payable"] == 13000
    overpaid = advance_summary(20000, 18000, 3000)
    assert overpaid["total_advance"] == 21000
    assert overpaid["balance_payable"] == -1000


def test_archive_hides_without_deleting(tmp_path):
    store = RequestStore(f"sqlite:///{tmp_path / 'archive.db'}")
    store.create({"trip_date": dt.date(2026, 7, 1), "vehicle_number": "MH01AA0001"})
    assert store.archive_before(dt.date(2026, 8, 1)) == 1
    assert store.list() == []
    assert len(store.list(include_archived=True)) == 1


def test_rtgs_roundtrip_filter_and_exact_export_columns(tmp_path):
    store = RequestStore(f"sqlite:///{tmp_path / 'rtgs.db'}")
    store.create({
        "report_scope": "RTGS", "trip_date": dt.date(2026, 8, 4), "vehicle_number": "",
        "beneficiary_name": "Demo Beneficiary", "amount": 4500, "status": "Verified",
        "rtgs_data": {"Pymt_Mode": "NEFT", "Beneficiary Account No": "001234",
                      "Bene_IFSC_Code": "DEMO0001234", "Remark": "Office expense",
                      "Pymt_Date": dt.date(2026, 8, 4)},
    })
    assert store.list(report_kind="DTR") == []
    rows = store.list(report_kind="RTGS")
    frame = rows_to_rtgs(rows)
    assert list(frame.columns) == RTGS_COLUMNS
    assert frame.iloc[0]["BNF_NAME"] == "Demo Beneficiary"
    assert frame.iloc[0]["BENE_ACC_NO"] == "001234"
    import xlrd
    ws = xlrd.open_workbook(file_contents=export_rtgs(frame), formatting_info=True).sheet_by_name("Sheet1")
    assert ws.row_values(0) == RTGS_COLUMNS
    assert ws.cell_value(1, 4) == "001234"


def test_rtgs_business_rules_group_trips_and_select_email_and_mode():
    rows = [{"BNF_NAME": "Demo & Co.", "BENE_ACC_NO": "00-123", "BENE_IFSC": "ICIC0000258",
             "AMOUNT": 4500, "Transporter Freight": 6000,
             "REMARK": f"7872 Talegaon to Pune 10MT 0{day} 08 2026 TA"} for day in (2, 3)]
    result = normalize_rtgs_records(rows, dt.date(2026, 8, 8))
    assert len(result) == 1
    assert result[0]["AMOUNT"] == 9000 and result[0]["Transporter Freight"] == 12000
    assert result[0]["REMARK"] == "7872 02 08 2026 to 03 08 2026 2Trp TA"
    assert result[0]["PYMT_MODE"] == "FT" and result[0]["EMAIL_ID"] == "jhanitish942@gmail.com"
    assert result[0]["MOBILE_NUM"] == "9028703567"
    assert result[0]["BNF_NAME"] == "Demo Co" and result[0]["BENE_ACC_NO"] == "00123"


def test_both_scope_appears_in_both_reports(tmp_path):
    store = RequestStore(f"sqlite:///{tmp_path / 'both.db'}")
    store.create({"report_scope": "Both", "trip_date": dt.date(2026, 8, 4),
                  "vehicle_number": "MH01AA0001", "rtgs_data": {"Remark": "Trip advance"}})
    assert len(store.list(report_kind="DTR")) == 1
    assert len(store.list(report_kind="RTGS")) == 1


def test_store_list_accepts_report_kind_after_interface_upgrade(tmp_path):
    store = RequestStore(f"sqlite:///{tmp_path / 'interface.db'}")
    assert store.list(report_kind="DTR") == []


def test_ai_dtr_result_maps_to_review_columns_without_inventing_values():
    result = DTRIntakeResult(rows=[DTRIntakeRow(
        date="2026-08-04", vehicle_number="MH14JL9818", lr_number="00127",
        diesel_quantity=None, review_notes="Freight is not visible",
    )])
    record = result_to_records("DTR", result)[0]
    assert record["Vehicle No."] == "MH14JL9818"
    assert record["LR No."] == "00127"
    assert record["Diesel Qty"] is None
    assert record["Review Notes"] == "Freight is not visible"


def test_intake_batch_keeps_multiple_source_files(tmp_path):
    store = RequestStore(f"sqlite:///{tmp_path / 'batch.db'}")
    batch_id = store.create_batch("DTR", "Shyam", "Two trips", [
        {"filename": "one.jpg", "mime_type": "image/jpeg", "data": b"one"},
        {"filename": "two.pdf", "mime_type": "application/pdf", "data": b"two"},
    ])
    attachments = store.get_attachments(batch_id)
    assert [item["filename"] for item in attachments] == ["one.jpg", "two.pdf"]
    assert [item["payload"] for item in attachments] == [b"one", b"two"]


def test_batch_rows_are_created_together(tmp_path):
    store = RequestStore(f"sqlite:///{tmp_path / 'many.db'}")
    numbers = store.create_many([
        {"trip_date": dt.date(2026, 8, 4), "vehicle_number": "MH01AA0001"},
        {"trip_date": dt.date(2026, 8, 4), "vehicle_number": "MH01AA0002"},
    ])
    assert len(numbers) == 2 and len(set(numbers)) == 2
    assert len(store.list()) == 2


def test_specialized_prompts_encode_real_whatsapp_workflow():
    dtr_prompt = _prompt("DTR", "two trips")
    rtgs_prompt = _prompt("RTGS", "one combined transfer")
    assert "separate DTR rows" in dtr_prompt
    assert "batch-total check" in dtr_prompt
    assert "Group consecutive trip blocks" in rtgs_prompt
    assert "photo-then-message sequence" in rtgs_prompt
    assert "belongs in rtgs_advance" in dtr_prompt
    assert "MICR" in dtr_prompt and "UPI QR requires verification" in rtgs_prompt


def test_only_model_availability_errors_trigger_fallback():
    assert _model_unavailable(Exception("404 NOT_FOUND: model is no longer available"))
    assert not _model_unavailable(Exception("429 RESOURCE_EXHAUSTED: quota exceeded"))
    assert not _model_unavailable(Exception("401 UNAUTHENTICATED: invalid API key"))


def test_ai_extraction_ignores_unapproved_expensive_model_override(monkeypatch):
    from google import genai

    calls = []

    class FakeModels:
        def generate_content(self, model, contents, config):
            calls.append(model)
            return type("Response", (), {"text": '{"rows": [], "summary": "No rows"}'})()

    class FakeClient:
        def __init__(self, api_key):
            self.models = FakeModels()

    monkeypatch.setattr(genai, "Client", FakeClient)
    result, model = extract_intake("test-key", "RTGS", "test prompt", [], model="retired-model")
    assert result.summary == "No rows"
    assert model == "gemini-3.1-flash-lite"
    assert calls == ["gemini-3.1-flash-lite"]


def test_ai_extraction_uses_only_flash_lite_fallback(monkeypatch):
    from google import genai

    calls = []

    class FakeModels:
        def generate_content(self, model, contents, config):
            calls.append(model)
            if model == "gemini-3.1-flash-lite":
                raise Exception("404 NOT_FOUND: model is unavailable")
            assert config.thinking_config.thinking_budget == 0
            assert str(config.media_resolution).endswith("MEDIA_RESOLUTION_MEDIUM")
            return type("Response", (), {"text": '{"rows": [], "summary": "Fallback"}'})()

    class FakeClient:
        def __init__(self, api_key):
            self.models = FakeModels()

    monkeypatch.setattr(genai, "Client", FakeClient)
    result, model = extract_intake("test-key", "RTGS", "test prompt", [])
    assert result.summary == "Fallback"
    assert model == "gemini-2.5-flash-lite"
    assert calls == ["gemini-3.1-flash-lite", "gemini-2.5-flash-lite"]


def test_batch_sync_retains_removed_rows_as_cancelled(tmp_path):
    store = RequestStore(f"sqlite:///{tmp_path / 'sync.db'}")
    batch_id = store.create_batch("RTGS", "Nikhat", "", [])
    base = {"trip_date": dt.date(2026, 8, 5), "vehicle_number": "", "status": "Verified"}
    store.sync_batch_records(batch_id, "RTGS", [
        {**base, "amount": 1000, "rtgs_data": {"Amount": 1000}},
        {**base, "amount": 2000, "rtgs_data": {"Amount": 2000}},
    ], "Nikhat", "initial_review")
    store.sync_batch_records(batch_id, "RTGS", [
        {**base, "amount": 1500, "rtgs_data": {"Amount": 1500}},
    ], "Nikhat", "manual_edit")
    active = store.get_batch_requests(batch_id, "RTGS")
    all_rows = store.get_batch_requests(batch_id, "RTGS", include_cancelled=True)
    assert len(active) == 1 and float(active[0]["amount"]) == 1500
    assert len(all_rows) == 2 and all_rows[1]["status"] == "Cancelled"


def test_explicit_batch_delete_removes_rows_and_attachments(tmp_path):
    store = RequestStore(f"sqlite:///{tmp_path / 'delete.db'}")
    batch_id = store.create_batch("RTGS", "Nikhat", "", [
        {"filename": "proof.jpg", "mime_type": "image/jpeg", "data": b"proof"},
    ])
    store.sync_batch_records(batch_id, "RTGS", [
        {"trip_date": dt.date(2026, 8, 5), "vehicle_number": "", "status": "Verified"},
    ])
    result = store.delete_batch(batch_id)
    assert result == {"requests": 1, "batches": 1}
    assert store.get_batch(batch_id) is None and store.get_attachments(batch_id) == []


def test_unified_record_can_be_edited_and_deleted_with_revision_history(tmp_path):
    store = WorkflowRequestStore(f"sqlite:///{tmp_path / 'unified.db'}")
    number = store.create({
        "report_scope": "Both", "trip_date": dt.date(2026, 8, 19),
        "vehicle_number": "MH14AB1234", "branch": "Pune", "status": "Submitted",
        "source_filename": "proof.jpg", "source_mime_type": "image/jpeg", "source_image": b"proof",
    })
    store.update(number, {"branch": "Wada"}, "records_tab", "Tester")
    assert store.get(number)["branch"] == "Wada"
    assert store.delete_request(number) == 1
    assert store.get(number) is None


def test_rtgs_done_status_is_persistent(tmp_path):
    store = WorkflowRequestStore(f"sqlite:///{tmp_path / 'rtgs-status.db'}")
    first = store.create({
        "report_scope": "Both", "trip_date": dt.date(2026, 9, 2),
        "vehicle_number": "MH14AB1234", "rtgs_advance": 1200,
    })
    second = store.create({
        "report_scope": "Both", "trip_date": dt.date(2026, 9, 2),
        "vehicle_number": "MH14AB5678", "rtgs_advance": 800,
    })
    assert store.get(first)["rtgs_done"] is False
    assert store.mark_rtgs_done([first]) == 1
    assert store.get(first)["rtgs_done"] is True
    assert store.get(second)["rtgs_done"] is False


def test_activity_log_is_persistent_and_newest_first(tmp_path):
    database = f"sqlite:///{tmp_path / 'activity.db'}"
    store = WorkflowRequestStore(database)
    store.log_action("Nitish", "Created trip record", "REQ-1", "Request - 23/08/26")
    store.log_action("Ashok", "Downloaded DTR report", details="01/08/2026 to 31/08/2026")

    logs = WorkflowRequestStore(database).list_activity_logs()
    assert [row["user_name"] for row in logs] == ["Ashok", "Nitish"]
    assert logs[0]["action"] == "Downloaded DTR report"
    assert logs[1]["request_number"] == "REQ-1"


def test_pnl_uses_trip_margin_and_direct_expense_categories():
    trips = [{"revenue": 100000, "transporter_freight": 70000, "branch": "Pune", "diesel_advance": 2500,
              "dtr_data": '{"Toll Expense": 500}'}]
    expenses = [{"categories": {"Salary": 5000, "Rent": 3000}}]
    rows = pnl_summary(trips, expenses)
    values = {row["Particular"]: row["Amount"] for row in rows}
    assert values["Gross Contribution"] == 30000
    assert values["Branch"] == "Pune"
    assert values["Toll charges"] == -500
    assert values["Diesel"] == -2500
    assert values["Additional expenses"] == -8000
    assert values["Total Direct Expenses"] == -8000
    assert values["Net Profit / (Loss)"] == 19000
    workbook = load_workbook(BytesIO(export_pnl(trips, expenses, dt.date(2026, 8, 1), dt.date(2026, 8, 31))))
    assert workbook["P&L"]["A1"].value.startswith("Profit & Loss")
    assert len(DIRECT_EXPENSE_COLUMNS) == 13


def test_pnl_includes_manish_passing_expense():
    rows = pnl_summary([], [{"categories": {"Passing expense": 1750}}])
    values = {row["Particular"]: row["Amount"] for row in rows}
    assert values["Passing expense"] == -1750
    assert values["Total Direct Expenses"] == -1750
    assert values["Net Profit / (Loss)"] == -1750


def test_own_vehicle_pnl_uses_requested_expenses():
    trips = [{
        "ownership_type": "Own", "revenue": 50000, "upi": 2500, "diesel_advance": 8000,
        "dtr_data": '{"Toll Expense": 1000, "Repairs & Maintenance": 1500}',
    }]
    expenses = [{"categories": {
        "Driver's salary": 5000, "EMI": 3000, "Insurance": 2000, "Vehicle Tax": 500,
    }}]
    rows = vehicle_pnl_summary(trips, expenses, "Own")
    values = {row["Particular"]: row["Amount"] for row in rows}
    assert list(values) == [
        "Revenue freight", "Route expenses (UPI)", "Toll charges", "Diesel amount",
        "Driver's salary", "EMI", "Insurance", "Vehicle Tax",
        "Repair and maintenance", "Net Profit / (Loss)",
    ]
    assert values["Route expenses (UPI)"] == -2500
    assert values["Repair and maintenance"] == -1500
    assert values["Net Profit / (Loss)"] == 26500


def test_own_vehicle_pnl_recovers_upi_from_embedded_dtr_data():
    trips = [{
        "ownership_type": "Own", "revenue": 10000, "upi": 0,
        "dtr_data": '{"UPI": 2750}',
    }]
    values = {row["Particular"]: row["Amount"] for row in vehicle_pnl_summary(trips, [], "Own")}
    assert values["Route expenses (UPI)"] == -2750
    assert values["Net Profit / (Loss)"] == 7250


def test_outside_vehicle_pnl_uses_transporter_and_additional_expenses():
    trips = [{"ownership_type": "Outside", "revenue": 50000, "transporter_freight": 35000}]
    expenses = [{"amount": 2000, "categories": {"Office & General expenses": 2000}}]
    rows = vehicle_pnl_summary(trips, expenses, "Outside")
    assert rows == [
        {"Particular": "Revenue", "Amount": 50000},
        {"Particular": "Transporter Freight", "Amount": -35000},
        {"Particular": "Additional expenses", "Amount": -2000},
        {"Particular": "Net Profit / (Loss)", "Amount": 13000},
    ]


def test_own_and_outside_pnl_are_horizontal_and_branch_wise():
    trips = [
        {"branch": "Pune", "ownership_type": "Own", "vehicle_number": "OWN-1", "revenue": 10000, "upi": 1000},
        {"branch": "Wada", "ownership_type": "Own", "vehicle_number": "OWN-2", "revenue": 8000, "diesel_advance": 2000},
        {"branch": "Pune", "ownership_type": "Outside", "vehicle_number": "OUT-1", "revenue": 20000, "transporter_freight": 15000},
    ]
    expenses = [
        {"vehicle_number": "OWN-1", "categories": {"Driver's salary": 500}},
        {"vehicle_number": "OUT-1", "amount": 750, "categories": {"Extra Expense": 750}},
    ]
    own = branch_vehicle_pnl_summary(trips, expenses, "Own")
    outside = branch_vehicle_pnl_summary(trips, expenses, "Outside")
    assert [row["Branch"] for row in own] == ["Pune", "Wada", "Total"]
    assert own[0]["Revenue freight"] == 10000 and own[0]["Driver's salary"] == -500
    assert own[-1]["Net Profit / (Loss)"] == 14500
    assert [row["Branch"] for row in outside] == ["Pune", "Total"]
    assert outside[0]["Transporter Freight"] == -15000
    assert outside[0]["Additional expenses"] == -750


def test_both_vehicle_pnl_is_horizontal_and_branch_wise():
    trips = [
        {"branch": "Pune", "ownership_type": "Own", "vehicle_number": "OWN-1", "revenue": 20000,
         "upi": 1000, "diesel_advance": 4000, "dtr_data": '{"Toll Expense": 500}'},
        {"branch": "Wada", "ownership_type": "Outside", "vehicle_number": "OUT-1", "revenue": 30000,
         "transporter_freight": 22000},
    ]
    expenses = [{
        "vehicle_number": "OWN-1", "amount": 2300,
        "categories": {"Driver's salary": 2000, "Extra Expense": 300},
    }]
    rows = branch_pnl_summary(trips, expenses)
    assert list(rows[0]) == BRANCH_PNL_COLUMNS
    assert [row["Branch"] for row in rows] == ["Pune", "Wada", "Total"]
    pune, wada, total = rows
    assert (pune["Revenue-Own"], pune["UPI"], pune["Diesel"], pune["Toll"]) == (20000, 1000, 4000, 500)
    assert pune["Driver's Salary"] == 2000
    assert pune["Extra Exp"] == 300
    assert (pune["Expense"], pune["Profit"]) == (7800, 12200)
    assert (wada["Revenue OS"], wada["Transporter Freight"], wada["Profit"]) == (30000, 22000, 8000)
    assert (total["Total Revenue"], total["Expense"], total["Profit"]) == (50000, 29800, 20200)


def test_business_memory_uses_repeated_verified_records_without_guessing():
    rows = [
        {"status": "Verified", "report_scope": "Both", "vehicle_number": "MH14JL9818",
         "vehicle_type": "10MT", "transporter_name": "XYZ", "company_name": "SG", "branch": "Pune",
         "beneficiary_name": "Demo", "rtgs_data": '{"BENE_ACC_NO":"00123","BENE_IFSC":"ICIC0001"}'},
        {"status": "Verified", "report_scope": "Both", "vehicle_number": "MH 14 JL 9818",
         "vehicle_type": "10MT", "transporter_name": "XYZ", "company_name": "SG", "branch": "Pune",
         "beneficiary_name": "Demo", "rtgs_data": '{"BENE_ACC_NO":"00123","BENE_IFSC":"ICIC0001"}'},
        {"status": "Cancelled", "report_scope": "Both", "vehicle_number": "MH14JL9818",
         "vehicle_type": "99MT", "transporter_name": "Wrong"},
    ]
    memory = build_business_memory(rows)
    vehicle = recall(memory, "vehicles", "mh-14-jl-9818")
    assert vehicle["vehicle_capacity"] == ("10MT", 2)
    assert vehicle["transporter_name"] == ("XYZ", 2)
    assert recall(memory, "companies", "sg")["branch"] == ("Pune", 2)
    assert recall(memory, "beneficiaries", "demo")["account_number"] == ("00123", 2)
    assert recall(memory, "vehicles", "UNKNOWN") == {}


def test_operational_dtr_export_uses_full_reference_shape():
    frame = pd.DataFrame([{column: "" for column in OPERATIONAL_DTR_COLUMNS}])
    frame.loc[0, "LR No."] = "00127"
    ws = load_workbook(BytesIO(export_operational_dtr(frame)))["DTR"]
    headers = [cell.value for cell in ws[1]]
    assert len(headers) == 35 and "Company Name" in headers and "Compnay Name" not in headers
    assert "LR No." in headers and "UPI " in headers
    assert ws["I2"].value == "00127" and ws["I2"].number_format == "@"


def test_hot_deploy_modules_expose_current_workflow_contract():
    assert callable(workflow_convert_rtgs_to_dtr)
    assert hasattr(WorkflowRequestStore, "list_batches")
    assert hasattr(WorkflowRequestStore, "sync_batch_records")
    assert hasattr(WorkflowRequestStore, "delete_batch")
    assert callable(workflow_branch_vehicle_pnl_summary)
