from io import BytesIO
import datetime as dt
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .dtr_generator import DTR_COLUMNS, FINANCIAL_COLUMNS


def _display_length(value):
    if value is None or (not isinstance(value, str) and pd.isna(value)): return 0
    if isinstance(value, (pd.Timestamp, dt.datetime, dt.date)): return 10
    return len(str(value))


def export_dtr(df, preserve_financials=False):
    output = BytesIO()
    safe = df.reindex(columns=DTR_COLUMNS).copy()
    if not preserve_financials:
        for col in FINANCIAL_COLUMNS: safe[col] = None
    safe["Date"] = pd.to_datetime(safe["Date"], errors="coerce").dt.date
    with pd.ExcelWriter(output, engine="openpyxl", date_format="DD-MM-YYYY") as writer:
        safe.to_excel(writer, index=False, sheet_name="DTR")
        ws = writer.book["DTR"]
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True); cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for index, name in enumerate(DTR_COLUMNS, 1):
            letter = get_column_letter(index)
            lengths = [_display_length(name)] + [_display_length(v) for v in safe[name].tolist()]
            width = min(max(max(lengths, default=0) + 2, 10), 32)
            if name == "Date": width = max(14, width)
            ws.column_dimensions[letter].width = width
            for cell in ws[letter][1:]:
                if name == "Date" and cell.value is not None: cell.number_format = "dd-mm-yyyy"
                elif name in ("Vehicle No.", "Invoice No."): cell.number_format = "@"
    return output.getvalue()
